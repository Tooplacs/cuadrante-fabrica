from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar
from datetime import date, timedelta

from schedule.models import ShiftAssignment
from employees.models import Employee

COLOR_HEADER_BG = '009B9B'
COLOR_NAME_BG   = '009B9B'
COLOR_TM_BG     = 'C6EFCE'
COLOR_TT_BG     = 'FFEB9C'
COLOR_TN_BG     = '0070C0'
COLOR_DF_BG     = 'FF0000'
COLOR_TEXT_TM   = '006100'
COLOR_TEXT_TT   = '9C5700'
COLOR_TEXT_TN   = '9C0006'
COLOR_TEXT_DF   = 'FFFFFF'
COLOR_WHITE     = 'FFFFFF'
COLOR_BLACK     = '000000'

SHIFT_HOURS = {
    'TM': '7:00',
    'TT': '15:00',
    'TN': '23:00',
    'DF': '0:00',
}


def get_border():
    thin = Side(style='thin', color='AAAAAA')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_cell(ws, row, col, value, bg_color, font_color=COLOR_BLACK,
               bold=False, italic=False, align='center', size=11):
    cell           = ws.cell(row=row, column=col, value=value)
    cell.fill      = PatternFill('solid', start_color=bg_color)
    cell.font      = Font(name='Calibri', bold=bold, italic=italic,
                          color=font_color, size=size)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border    = get_border()
    return cell


def build_excel(year, start_month=1, num_months=12):
    employees = list(Employee.objects.filter(is_active=True).order_by('name'))
    wb        = Workbook()

    # ═══════════════════════════════
    # ONGLET 1 : Cuadrante mensuel
    # ═══════════════════════════════
    ws       = wb.active
    ws.title = f'Cuadrante {year}'

    for i in range(1, len(employees) + 6):
        ws.row_dimensions[i].height = 20
    ws.column_dimensions['A'].width = 22

    months = []
    m_year, m_month = year, start_month
    for _ in range(num_months):
        abbr  = calendar.month_abbr[m_month]
        label = f"{abbr}-{str(m_year)[2:]}"
        months.append((m_year, m_month, label))
        if m_month == 12:
            m_year, m_month = m_year + 1, 1
        else:
            m_month += 1

    for i in range(num_months):
        ws.column_dimensions[get_column_letter(i + 2)].width = 8

    apply_cell(ws, 1, 1, '', COLOR_HEADER_BG, COLOR_WHITE, bold=True)
    for col_idx, (m_year, m_month, label) in enumerate(months, start=2):
        apply_cell(ws, 1, col_idx, label, COLOR_HEADER_BG, COLOR_WHITE, bold=True)

    for row_idx, emp in enumerate(employees, start=2):
        apply_cell(ws, row_idx, 1, emp.name, COLOR_NAME_BG,
                   COLOR_WHITE, bold=True, italic=True, align='left')
        for col_idx, (m_year, m_month, label) in enumerate(months, start=2):
            try:
                assignment = ShiftAssignment.objects.get(
                    employee=emp, year=m_year, month=m_month)
                shift = assignment.shift
            except ShiftAssignment.DoesNotExist:
                shift = ''
            if shift == 'TM':
                bg, fg = COLOR_TM_BG, COLOR_TEXT_TM
            elif shift == 'TT':
                bg, fg = COLOR_TT_BG, COLOR_TEXT_TT
            elif shift == 'TN':
                bg, fg = COLOR_TN_BG, COLOR_TEXT_TN
            else:
                bg, fg = COLOR_WHITE, COLOR_BLACK
            apply_cell(ws, row_idx, col_idx, shift, bg, fg, bold=True)

    count_row_tm = len(employees) + 2
    count_row_tt = len(employees) + 3
    count_row_tn = len(employees) + 4

    apply_cell(ws, count_row_tm, 1, 'TM', COLOR_TM_BG, COLOR_TEXT_TM, bold=True, italic=True)
    apply_cell(ws, count_row_tt, 1, 'TT', COLOR_TT_BG, COLOR_TEXT_TT, bold=True, italic=True)
    apply_cell(ws, count_row_tn, 1, 'TN', COLOR_TN_BG, COLOR_TEXT_TN, bold=True, italic=True)

    for col_idx, (m_year, m_month, label) in enumerate(months, start=2):
        col_letter    = get_column_letter(col_idx)
        first_emp_row = 2
        last_emp_row  = len(employees) + 1
        apply_cell(ws, count_row_tm, col_idx,
                   f'=COUNTIF({col_letter}{first_emp_row}:{col_letter}{last_emp_row},"TM")',
                   COLOR_WHITE, COLOR_BLACK, bold=True)
        apply_cell(ws, count_row_tt, col_idx,
                   f'=COUNTIF({col_letter}{first_emp_row}:{col_letter}{last_emp_row},"TT")',
                   COLOR_WHITE, COLOR_BLACK, bold=True)
        apply_cell(ws, count_row_tn, col_idx,
                   f'=COUNTIF({col_letter}{first_emp_row}:{col_letter}{last_emp_row},"TN")',
                   COLOR_WHITE, COLOR_BLACK, bold=True)

    # ═══════════════════════════════
    # ONGLET 2 : Cuadrante RRHH
    # ═══════════════════════════════
    ws2        = wb.create_sheet(title=f'Cuadrante RRHH {year}')
    n_emp      = len(employees)

    # Tous les jours de l'annee
    days = []
    d    = date(year, 1, 1)
    while d.year == year:
        days.append(d)
        d += timedelta(days=1)

    # Largeurs colonnes
    ws2.column_dimensions['A'].width = 25
    for day_idx in range(len(days)):
        col_code  = 2 + day_idx * 2
        col_heure = col_code + 1
        ws2.column_dimensions[get_column_letter(col_code)].width  = 4
        ws2.column_dimensions[get_column_letter(col_heure)].width = 6

    # Hauteurs lignes
    ws2.row_dimensions[1].height = 20
    ws2.row_dimensions[2].height = 20
    for i in range(3, n_emp + 6):
        ws2.row_dimensions[i].height = 18

    # Ligne 1 : dates
    apply_cell(ws2, 1, 1, '', COLOR_HEADER_BG, COLOR_WHITE, bold=True, size=8)
    apply_cell(ws2, 2, 1, '', COLOR_HEADER_BG, COLOR_WHITE, bold=True, size=8)

    for day_idx, day in enumerate(days):
        col_code   = 2 + day_idx * 2
        col_heure  = col_code + 1
        is_weekend = day.weekday() >= 5
        header_bg  = 'CC0000' if is_weekend else COLOR_HEADER_BG
        label      = day.strftime('%d/%m/%Y')

        apply_cell(ws2, 1, col_code,  label,   header_bg, COLOR_WHITE, bold=True, size=7)
        apply_cell(ws2, 1, col_heure, '',       header_bg, COLOR_WHITE, bold=True, size=7)
        apply_cell(ws2, 2, col_code,  'Turno',  header_bg, COLOR_WHITE, bold=True, size=7)
        apply_cell(ws2, 2, col_heure, 'Hora',   header_bg, COLOR_WHITE, bold=True, size=7)

    # Recuperer tous les quarts en une seule requete
    all_assignments = ShiftAssignment.objects.filter(year=year).select_related('employee')
    shift_map = {}
    for a in all_assignments:
        shift_map[(a.employee.id, a.month)] = a.shift

    # Lignes employes
    for row_idx, emp in enumerate(employees, start=3):
        apply_cell(ws2, row_idx, 1, emp.name, COLOR_NAME_BG,
                   COLOR_WHITE, bold=True, italic=True, align='left', size=9)

        for day_idx, day in enumerate(days):
            col_code   = 2 + day_idx * 2
            col_heure  = col_code + 1
            is_weekend = day.weekday() >= 5

            if is_weekend:
                shift = 'DF'
                bg    = COLOR_DF_BG
                fg    = COLOR_TEXT_DF
            else:
                shift = shift_map.get((emp.id, day.month), '')
                if shift == 'TM':
                    bg, fg = COLOR_TM_BG, COLOR_TEXT_TM
                elif shift == 'TT':
                    bg, fg = COLOR_TT_BG, COLOR_TEXT_TT
                elif shift == 'TN':
                    bg, fg = COLOR_TN_BG, COLOR_TEXT_TN
                else:
                    bg, fg = COLOR_WHITE, COLOR_BLACK

            heure = SHIFT_HOURS.get(shift, '')
            apply_cell(ws2, row_idx, col_code,  shift, bg, fg, bold=True, size=9)
            apply_cell(ws2, row_idx, col_heure, heure, bg, fg, bold=True, size=9)

    # Lignes compteurs RRHH
    count_start = n_emp + 3
    apply_cell(ws2, count_start,     1, 'MAÑANA', COLOR_TM_BG, COLOR_TEXT_TM, bold=True, size=9)
    apply_cell(ws2, count_start + 1, 1, 'TARDE',  COLOR_TT_BG, COLOR_TEXT_TT, bold=True, size=9)
    apply_cell(ws2, count_start + 2, 1, 'NOCHE',  COLOR_TN_BG, COLOR_TEXT_TN, bold=True, size=9)

    for day_idx, day in enumerate(days):
        col_code      = 2 + day_idx * 2
        col_heure     = col_code + 1
        col_letter    = get_column_letter(col_code)
        first_emp_row = 3
        last_emp_row  = n_emp + 2

        apply_cell(ws2, count_start,     col_code,
                   f'=COUNTIF({col_letter}{first_emp_row}:{col_letter}{last_emp_row},"TM")',
                   COLOR_WHITE, COLOR_BLACK, bold=True, size=9)
        apply_cell(ws2, count_start + 1, col_code,
                   f'=COUNTIF({col_letter}{first_emp_row}:{col_letter}{last_emp_row},"TT")',
                   COLOR_WHITE, COLOR_BLACK, bold=True, size=9)
        apply_cell(ws2, count_start + 2, col_code,
                   f'=COUNTIF({col_letter}{first_emp_row}:{col_letter}{last_emp_row},"TN")',
                   COLOR_WHITE, COLOR_BLACK, bold=True, size=9)

        apply_cell(ws2, count_start,     col_heure, '', COLOR_WHITE, COLOR_BLACK, size=9)
        apply_cell(ws2, count_start + 1, col_heure, '', COLOR_WHITE, COLOR_BLACK, size=9)
        apply_cell(ws2, count_start + 2, col_heure, '', COLOR_WHITE, COLOR_BLACK, size=9)

    return wb